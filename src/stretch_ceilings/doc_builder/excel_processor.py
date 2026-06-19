def _find_header_row(ws):
    for row in ws.iter_rows(min_row=3):
        non_empty = [c for c in row if c.value is not None]
        if len(non_empty) >= 2:
            return row[0].row
    return None


def _find_area_columns(ws, header_row_idx):
    return {
        str(cell.value).strip(): cell.column
        for cell in ws[header_row_idx]
        if cell.value and "area" in str(cell.value).lower()
    }


def _find_total_row(ws, header_row_idx):
    for row in ws.iter_rows(min_row=header_row_idx + 1):
        for cell in row:
            if cell.value and str(cell.value).strip().lower() == "total":
                return cell.row
    return None


def get_areas_data(areas_ws):
    result = {}

    result["sheet_grade"] = next(
        (cell.value for cell in areas_ws[2] if cell.value is not None), None
    )

    header_row_idx = _find_header_row(areas_ws)
    if header_row_idx is None:
        return result

    area_cols = _find_area_columns(areas_ws, header_row_idx)
    total_row_idx = _find_total_row(areas_ws, header_row_idx)

    if total_row_idx is None or not area_cols:
        return result

    if len(area_cols) == 1:
        col_idx = next(iter(area_cols.values()))
        val = areas_ws.cell(row=total_row_idx, column=col_idx).value
        result["chargeable_area"] = val
        result["actual_area"] = val
    else:
        for col_name, col_idx in area_cols.items():
            val = areas_ws.cell(row=total_row_idx, column=col_idx).value
            if "chargeable" in col_name.lower():
                result["chargeable_area"] = val
            elif "actual" in col_name.lower():
                result["actual_area"] = val

    return result


def get_profiles_data(profiles_ws):
    result = {}

    header_row_idx = _find_header_row(profiles_ws)
    if header_row_idx is None:
        return result

    # Map each header to its column index
    col_map = {
        str(cell.value).strip(): cell.column
        for cell in profiles_ws[header_row_idx]
        if cell.value is not None
    }

    # Grade column: first column in the header map
    # Length column: whichever header contains "length"
    grade_col = next(iter(col_map.values()), None)
    length_col = next(
        (col for header, col in col_map.items() if "length" in header.lower()), None
    )

    if grade_col is None or length_col is None:
        return result

    profiles = []
    for row in profiles_ws.iter_rows(min_row=header_row_idx + 1):
        grade_val = profiles_ws.cell(row=row[0].row, column=grade_col).value
        length_val = profiles_ws.cell(row=row[0].row, column=length_col).value
        if grade_val is not None:
            profiles.append({"grade": grade_val, "length": length_val})

    result["profiles"] = profiles
    return result


def get_lights_data(lights_ws):
    return {
        "light_type": lights_ws.cell(row=2, column=1).value,  # A2
        "box_depth":  lights_ws.cell(row=2, column=4).value,  # D2
    }


def get_area_data(wb):
    areas_ws = wb.worksheets[0]
    profiles_ws = wb.worksheets[1]

    data = get_areas_data(areas_ws)
    data.update(get_profiles_data(profiles_ws))

    try:
        data.update(get_lights_data(wb["Lights BOM"]))
    except KeyError:
        pass

    data["product"] = "Stretch Ceilings"
    return data


_PRODUCT_MODULES = {
    "plain translucent":   "plain_translucent",
    "printed translucent": "printed_translucent",
    "mirror":              "mirror",
    "lacquered":           "lacquered",
}


def product_convert(option):
    import importlib
    grade = str(option.get("sheet_grade") or "").strip().lower()
    module_name = next(
        (mod for key, mod in _PRODUCT_MODULES.items() if key in grade),
        "plain_translucent",
    )
    mod = importlib.import_module(f"stretch_ceilings.doc_builder.products.{module_name}")
    return mod.convert


def get_lights_commercials(data):
    import openpyxl
    import excel_utils
    from stretch_ceilings.doc_builder.products._base import read_lights_bom, generate_lights_df

    results = []
    for area in data["areas"]:
        for option in data["areas"][area]:
            wb_file = option.get("area_table")
            if wb_file is None:
                continue
            if hasattr(wb_file, "seek"):
                wb_file.seek(0)
            wb = openpyxl.load_workbook(wb_file, data_only=True)
            lights_bom = read_lights_bom(wb)
            if lights_bom:
                lights_df = generate_lights_df(lights_bom)
                commercial_xl = excel_utils.generate_commercial_table(lights_df, freight=False)
                results.append(("Lights & Drivers Commercials.xlsx", commercial_xl))

    return results
