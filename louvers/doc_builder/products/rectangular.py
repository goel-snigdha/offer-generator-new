import math
import openpyxl
from pathlib import Path
from louvers.doc_builder.excel_utils import FINISH_RATE_COLS
from excel_utils import (
    get_total_cols,
    get_max_row,
    set_cell_and_save,
    get_cell_ref,
    evaluate_formula,
)

END_CAP_RATE = 120
INSTALLATION_RATE = 200


def get_data(xl):

    vars = {}

    pitch = xl.cell(row=1, column=2).value
    vars["pitch"] = pitch

    section_type = xl.cell(row=1, column=3).value
    vars["section_type"] = section_type.replace("Rectangular Louvers ", "")

    return vars


def update_data(xl, curr_row):

    vars = get_data(xl)
    vars.update(get_total_cols(xl, curr_row, 6, 11))

    return vars


def generate_df(window_data, finish, installation, rate):

    string = "Supply of Rectangular Louvers {} in {} Finish at Pitch {}mm"
    offer_rate_formula = f"=CEILING({rate}*1.01, 10)"
    area = math.ceil(float(window_data["Area (ft2)"]))

    data = [
        [
            string.format(window_data["section_type"], finish, window_data["pitch"]),
            area,
            "ft\u00b2",
            offer_rate_formula,
        ]
    ]

    if "End Caps (pcs)" in window_data:
        if window_data["End Caps (pcs)"] > 0:
            qty = round(window_data["End Caps (pcs)"], 0)
            data.append(["End Caps", qty, "pcs", END_CAP_RATE])

    if installation:
        data.append(["Installation Charges", area, "ft\u00b2", INSTALLATION_RATE])

    return data


def convert(window_wb, data, installation):

    finish = data['finish']

    window_xl = window_wb.worksheets[0]
    max_row = get_max_row(window_xl)

    vars = update_data(window_xl, max_row)
    BASE_DIR = Path(__file__).resolve().parents[2]
    ext = f'rectangular_{ vars["section_type"].replace("x", "_")}.xlsx'
    path = BASE_DIR/"files"/"reference_xls"/"price_xls"/ext
    price_wb = openpyxl.load_workbook(path, data_only=False)

    wb_with_pitch = set_cell_and_save(price_wb, vars["pitch"], 7, 2)

    cell_ref = get_cell_ref(FINISH_RATE_COLS[finish], 4)
    rate = evaluate_formula(wb_with_pitch, "Price", cell_ref)
    data = generate_df(vars, finish, installation, rate)

    return data
