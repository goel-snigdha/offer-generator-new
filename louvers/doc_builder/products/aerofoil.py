import openpyxl
from pathlib import Path
from louvers.doc_builder.excel_utils import FINISH_RATE_COLS
from excel_utils import (
    get_total_cols,
    get_max_row,
    set_cell_and_save,
    get_cell_ref,
    evaluate_formula,
)

INSTALLATION_RATE = {
    "C-Channel": 200,
    "Fringe End Caps": 200,
    "D-Wall Bracket": 200,
    "Slot Cut Pipe": 200,
    "Manually Moveable": 300,
    "Motorized": 200,
}
MECHANISM_ACCESSORIES = {
    "C-Channel": ("Supply of C-Channel", "Total C-Channel Length (m)", "m", 750),
    "Fringe End Caps": ("Supply of End Caps", "Fringe End Caps (pcs)", "pcs", 250),
    "D-Wall Bracket": ("Supply of End Caps", "Total End Caps (pcs)", "pcs", 250),
    "Slot Cut Pipe": ("Slot Cut Pipe Length", "Total MS Rod Length (m)", "m", 0),
    "Manually Moveable": ("Manually Moveable Mechanism", "Area (ft2)", "ft\u00b2", 500),
    "Moveable Motorized": ("Motorization", "No. of Motors (pcs)", "pcs", 4200),
}


def get_data(xl):

    vars = {}

    pitch = xl.cell(row=1, column=2).value
    vars["pitch"] = pitch

    section_type = xl.cell(row=1, column=3).value
    vars["section_type"] = section_type.replace("Aerofoil ", "")

    for col in range(8, 12):
        if xl.cell(row=1, column=col).value == "Installation":
            installation_col = col
    installation_method = xl.cell(row=1, column=installation_col + 1).value
    vars["installation_method"] = installation_method

    return vars


def update_data(xl, curr_row):

    vars = get_data(xl)
    vars.update(get_total_cols(xl, curr_row, 6, 10))

    return vars


def generate_df(window_data, finish, installation, af_rate):

    offer_rate_formula = f"=CEILING({af_rate}*1.01, 10)"

    area = round(window_data["Area (ft2)"], 0)
    data = [
        [
            "Supply of Aerofoil-{} in {} Finish at Pitch {}mm".format(
                window_data["section_type"], finish, window_data["pitch"]
            ),
            area,
            "ft\u00b2",
            offer_rate_formula,
        ]
    ]

    if window_data["installation_method"] in MECHANISM_ACCESSORIES:
        title, qty_col, uom, rate = MECHANISM_ACCESSORIES[
            window_data["installation_method"]
        ]
        if qty_col in window_data:
            qty = round(window_data[qty_col], 0)
            if qty > 0:
                data.append([title, qty, uom, rate])

    installation = INSTALLATION_RATE[window_data["installation_method"]]
    if installation:
        data.append([
            "Installation Charges",
            area,
            "ft\u00b2",
            installation
        ])

    return data


def convert(window_wb, data, installation):

    finish = data['finish']

    window_xl = window_wb.worksheets[0]
    max_row = get_max_row(window_xl)
    vars = update_data(window_xl, max_row)

    BASE_DIR = Path(__file__).resolve().parents[2]
    ext = f'aerofoil_{vars["section_type"].lower()}.xlsx'
    path = BASE_DIR/"files"/"reference_xls"/"price_xls"/ext
    price_wb = openpyxl.load_workbook(
        path, data_only=False
    )

    pitch_row = 7 if vars["section_type"] == "AF400" else 8
    wb_with_pitch = set_cell_and_save(price_wb, vars["pitch"], pitch_row, 2)

    price_row = 4 if vars["section_type"] == "AF400" else 5
    cell_ref = get_cell_ref(FINISH_RATE_COLS[finish], price_row)
    rate = evaluate_formula(wb_with_pitch, "Price", cell_ref)
    data = generate_df(vars, finish, installation, rate)

    return data
