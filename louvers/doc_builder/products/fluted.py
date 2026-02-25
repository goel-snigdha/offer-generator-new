import openpyxl
from pathlib import Path
from excel_utils import get_total_cols, get_max_row
from louvers.doc_builder.excel_utils import FINISH_RATE_COLS

EPDM_M_RATE = 50
ALUMINUM_FRAME = 150
INSTALLATION_RATE = 200


def get_data(xl):

    vars = {}

    finish_sides = xl.cell(row=1, column=10).value
    vars["finish_sides"] = finish_sides

    return vars


def update_data(xl, curr_row):

    vars = get_data(xl)
    vars.update(get_total_cols(xl, curr_row, 6, 10))

    return vars


def generate_df(window_data, finish, installation, fluted_rate):

    offer_rate_formula = f"=CEILING({fluted_rate}*1.01, 10)"
    frame_rate = ALUMINUM_FRAME
    area = round(window_data["Area (ft2)"], 0)
    epdm = round(window_data["EPDM Gasket Length (m)"], 0)

    data = [
        [
            "Supply of Fluted Panels in {} Finish".format(finish),
            area,
            "ft\u00b2",
            offer_rate_formula,
        ],
        ["Supply of EPDM Rubber", epdm, "m", EPDM_M_RATE],
        ["Aluminium Framing", area, "ft\u00b2", frame_rate],
    ]

    if installation:
        data.append(["Installation Charges", area, "ft\u00b2", INSTALLATION_RATE])

    return data


def convert(window_wb, data, installation):

    finish = data['finish']

    window_xl = window_wb.worksheets[0]
    max_row = get_max_row(window_xl)
    vars = update_data(window_xl, max_row)

    BASE_DIR = Path(__file__).resolve().parents[2]
    path = BASE_DIR/"files"/"reference_xls"/"price_xls"/"fluted.xlsx"
    price_wb = openpyxl.load_workbook(path, data_only=True)
    price_xl = price_wb.worksheets[0]

    price_row = 4
    if vars["finish_sides"] == "Single":
        price_row = 3
    rate = price_xl.cell(row=price_row, column=FINISH_RATE_COLS[finish]).value

    data = generate_df(vars, finish, installation, rate)

    return data
