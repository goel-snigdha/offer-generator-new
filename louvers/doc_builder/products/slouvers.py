import openpyxl
from pathlib import Path
from excel_utils import get_total_cols, get_max_row
from louvers.doc_builder.excel_utils import FINISH_RATE_COLS

INSTALLATION_RATE = 200


def get_data(xl):

    vars = {}

    pitch = xl.cell(row=1, column=2).value
    vars["pitch"] = pitch

    finish_sides = xl.cell(row=1, column=9).value
    vars["finish_sides"] = finish_sides

    return vars


def update_data(xl, curr_row):

    vars = get_data(xl)
    vars.update(get_total_cols(xl, curr_row, 6, 11))

    return vars


def generate_df(window_data, finish, installation, louver_rate):

    offer_rate_formula = f"=CEILING({louver_rate}*1.01, 10)"
    area = round(window_data["Area (ft2)"], 0)

    data = [
        [
            "Supply of S-Louvers in {} Finish at Pitch {}mm".format(
                finish, window_data["pitch"]
            ),
            area,
            "ft\u00b2",
            offer_rate_formula,
        ]
    ]

    if installation:
        data.append(["Installation Charges", area, "ft\u00b2", INSTALLATION_RATE])

    return data


def convert(window_wb, data, installation):

    finish = data['finish']

    window_xl = window_wb.worksheets[0]
    max_row = get_max_row(window_xl)
    vars = update_data(window_xl, max_row)

    BASE_DIR = Path(__file__).resolve().parents[2]
    path = BASE_DIR/"files"/"reference_xls"/"price_xls"/"slouver.xlsx"
    price_wb = openpyxl.load_workbook(path, data_only=True)
    price_xl = price_wb.worksheets[0]

    price_row = 5 if vars["finish_sides"] == "Single" else 6
    rate = price_xl.cell(row=price_row, column=FINISH_RATE_COLS[finish]).value

    data = generate_df(vars, finish, installation, rate)

    return data
