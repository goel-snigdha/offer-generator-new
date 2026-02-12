import openpyxl
from doc_builder.excel_utils import get_orientation, generate_commercial_table, combine_xls, number_to_alpha
from doc_builder.products import (
    grille,
    cottal,
    fluted,
    aerofoil,
    slouvers,
    rectangular,
    beamc,
)

PRODUCT_FUNCTIONS = {
    "Grille": [grille.get_data, grille.convert],
    "Cottal": [cottal.get_data, cottal.convert],
    "Fluted": [fluted.get_data, fluted.convert],
    "Aerofoil": [aerofoil.get_data, aerofoil.convert],
    "S-Louvers": [slouvers.get_data, slouvers.convert],
    "Rectangular": [rectangular.get_data, rectangular.convert],
    "Beam C-Channel": [beamc.get_data, beamc.convert],
}


def get_area_data(xl):

    area_data = {}

    val = xl.cell(row=1, column=3).value
    for product in PRODUCT_FUNCTIONS.keys():
        if val and val.startswith(product):
            area_data['product'] = product

    if xl.cell(row=1, column=1).value.startswith("Beam C-Channel"):
        area_data['product'] = "Beam C-Channel"

    area_data['orientation'] = get_orientation(xl)
    get_data_func = PRODUCT_FUNCTIONS.get(area_data["product"], [])[0]
    new_area_data = get_data_func(xl)
    area_data.update(new_area_data)

    return area_data


def convert(data):

    output_dfs = []
    output_xls = []
    found_options = False

    for area in data["areas"]:
        if len(data["areas"][area]) > 1:
            found_options = True

        options = data["areas"][area]

        line_item_str = str(area)
        option_str = ""

        for idx in range(len(options)):
            option = data["areas"][area][idx]

            wb = option["area_table"]
            window_wb = openpyxl.load_workbook(wb, data_only=True)

            option_str = ""
            if len(options) > 1:
                option_str = "Option " + number_to_alpha(idx + 1)

            option["line_item_str"] = line_item_str
            option["option_str"] = option_str
            option_data_from_xl = get_area_data(window_wb.worksheets[0])
            option.update(option_data_from_xl)

            convert = PRODUCT_FUNCTIONS.get(option['product'], [])[1]

            if convert:
                output_df = convert(
                    window_wb, option["finish"], data['offer_data']["installation"]
                )
                output_dfs.append(output_df)

                line_item_title = option["product"] + " " + option["line_item_str"]
                line_item_title = line_item_title + option["option_str"].replace("Option ", "")
                filename = f"Commercials - {line_item_title}.xlsx"

                commercial_xl = generate_commercial_table(output_df)
                output_xls.append((filename, commercial_xl))
                option["CommercialTable"] = commercial_xl

            else:
                raise ValueError("This Excel could not be processed.")

    if not found_options:
        offer_num = data['offer_data']["OfferNumber"].replace("/", "-")
        combined_xl = (f"Commercials for {offer_num}.xlsx", combine_xls(output_dfs))
        output_xls.append(combined_xl)
        data["CombinedCommercials"] = combined_xl

    return output_xls, data
