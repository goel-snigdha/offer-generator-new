import openpyxl
from doc_builder.excel_utils import get_total_cols, get_max_row, FINISH_RATE_COLS

EPDM_M_RATE = 50
ALUMINUM_FRAME = 150
INSTALLATION_RATE = 200


def get_data(xl):

    vars = {}

    section_type = xl.cell(row=1, column=3).value
    vars["section_type"] = section_type.replace("Cottal ", "")

    return vars


def update_data(xl, curr_row):

    vars = get_data(xl)
    vars.update(get_total_cols(xl, curr_row, 6, 10))

    return vars


def generate_df(window_data, finish, installation, cottal_rate):

    offer_rate_formula = f"=CEILING({cottal_rate}*1.01, 10)"
    frame_rate = ALUMINUM_FRAME
    area = round(window_data["Area (ft2)"], 0)
    epdm = round(window_data["EPDM Gasket Length (m)"], 0)

    data = [
        [
            "Supply of Cottal in {} Finish".format(finish),
            area,
            "ft\u00b2",
            offer_rate_formula,
        ],
        ["Supply of EPDM Rubber", epdm, "m", EPDM_M_RATE],
        ["Aluminum Framing", area, "ft\u00b2", frame_rate],
    ]

    if installation:
        data.append(["Installation Charges", area, "ft\u00b2", INSTALLATION_RATE])

    return data


def convert(window_wb, finish, installation):

    window_xl = window_wb.worksheets[0]
    max_row = get_max_row(window_xl)
    vars = update_data(window_xl, max_row)

    path = "files/reference_xls/price_xls"
    ext = vars["section_type"].replace(" mm", "")
    price_wb = openpyxl.load_workbook(f"{path}/cottal_{ext}.xlsx", data_only=True)
    price_xl = price_wb.worksheets[0]

    rate = price_xl.cell(row=3, column=FINISH_RATE_COLS[finish]).value
    data = generate_df(vars, finish, installation, rate)

    return data
