import math
from collections import defaultdict
import openpyxl
from doc_builder.excel_utils import get_max_row

WIDTH_BAND_ROWS = [
    (0, 300),
    (300, 350),
    (350, 400),
    (400, 450),
    (450, 500),
    (500, 550),
    (550, 600),
    (600, 650),
    (650, 700),
    (700, 750),
    (750, 800),
]
INSTALLATION_COST = 300


def get_data(xl):

    return {}


def calc_price(xl, max_row):

    path = "files/reference_xls/price_xls"
    price_wb = openpyxl.load_workbook(f"{path}/beamc.xlsx", data_only=True)
    price_xl = price_wb.worksheets[0]
    vars = []
    unique_widths = defaultdict(float)

    for row in range(4, max_row):

        width = xl.cell(row=row, column=3).value
        total_length = xl.cell(row=row, column=6).value
        unique_widths[width] += total_length

    for width in unique_widths:

        price_row = 0
        for i in range(len(WIDTH_BAND_ROWS)):
            band = WIDTH_BAND_ROWS[i]
            if (width > band[0]) and (width <= band[1]):
                price_row = i + 2

        rate = price_xl.cell(row=price_row, column=2).value

        total_length = math.ceil(unique_widths[width])
        vars.append((width, total_length, rate))

    return vars


def generate_df(data):

    df = []
    string = "Supply of Beam C-Channel of {}'' in Powder Coated Finish"

    for beam in data:

        beam_rate_org = beam[2]
        beam_rate_formula = f"=CEILING({beam_rate_org}*1.01, 10)"
        length = round(beam[1], 0)

        df.append(
            [
                string.format(str(round(float(beam[0]) / 25.4))),
                length,
                "m",
                beam_rate_formula,
            ]
        )
        df.append(["Installation Charges", length, "m", INSTALLATION_COST])

    return df


def convert(window_wb, finish, installation):

    window_xl = window_wb.worksheets[0]
    max_row = get_max_row(window_xl)
    vars = calc_price(window_xl, max_row)
    data = generate_df(vars)

    return data
