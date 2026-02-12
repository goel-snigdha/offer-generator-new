import openpyxl
from doc_builder.excel_utils import (
    get_total_cols,
    get_max_row,
    set_cell_and_save,
    get_cell_ref,
    evaluate_formula,
    FINISH_RATE_COLS,
)

END_CAP_RATE = 50
JOINING_PC_RATE = 120
INSTALLATION_RATE = 200


def get_data(xl):

    vars = {}

    section_type = xl.cell(row=1, column=3).value
    vars["section_type"] = section_type.replace("Grille ", "")

    pitch = xl.cell(row=1, column=2).value
    vars["pitch"] = pitch

    return vars


def update_data(xl, curr_row):

    vars = get_data(xl)
    vars.update(get_total_cols(xl, curr_row, 6, 11))

    return vars


def generate_df(window_data, finish, installation, grille_rate):

    offer_rate_formula = f"=CEILING({grille_rate}*1.01, 10)"
    area = round(window_data["Area (ft2)"], 0)

    data = [
        [
            "Supply of Grilles 2550 in {} Finish at Pitch {}mm".format(
                finish, window_data["pitch"]
            ),
            area,
            "ft\u00b2",
            offer_rate_formula,
        ]
    ]

    if "End Caps (pcs)" in window_data:
        if window_data["End Caps (pcs)"] > 0:
            qty = round(window_data["End Caps (pcs)"], 0)
            data.append(["End Caps", qty, "pcs", END_CAP_RATE])
    if "Joining Pieces (pcs)" in window_data:
        if window_data["Joining Pieces (pcs)"] > 0:
            qty = round(window_data["Joining Pieces (pcs)"], 0)
            data.append(["Joining Pieces", qty, "pcs", JOINING_PC_RATE])

    if installation:
        data.append(["Installation Charges", area, "ft\u00b2", INSTALLATION_RATE])

    return data


def convert(window_wb, finish, installation):

    window_xl = window_wb.worksheets[0]
    max_row = get_max_row(window_xl)

    vars = update_data(window_xl, max_row)
    path = "files/reference_xls/price_xls"
    price_wb = openpyxl.load_workbook(f"{path}/grille.xlsx", data_only=False)

    wb_with_pitch = set_cell_and_save(price_wb, vars["pitch"], 8, 2)

    cell_ref = get_cell_ref(FINISH_RATE_COLS[finish], 5)
    rate = evaluate_formula(wb_with_pitch, "Price", cell_ref)
    data = generate_df(vars, finish, installation, rate)

    return data
