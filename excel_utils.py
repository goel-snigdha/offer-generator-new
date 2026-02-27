import math
import openpyxl
import streamlit as st
from pathlib import Path
from io import BytesIO
from xlcalculator import ModelCompiler, Evaluator
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

COLORS = ["F2F2F2", "FFFFFF"]


def number_to_alpha(n):
    if 1 <= n <= 26:
        return chr(64 + n)
    else:
        raise ValueError("Number must be between 1 and 26")


def merge_cells(xl, row_start, row_end, col_start, col_end):
    xl.merge_cells(
        start_row=row_start, start_column=col_start, end_row=row_end, end_column=col_end
    )


def format(cell):
    cell.font = Font(name="Montserrat", bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    return cell


def set_cell(cell, value, alignment="center", bold=False):
    cell.value = value
    cell.font = Font(name="Montserrat", bold=bold)
    cell.alignment = Alignment(horizontal=alignment, vertical="center")


def get_orientation(ws, start_row=4, col=3):

    orientations = set()
    row = start_row

    while True:
        cell_value = ws.cell(row=row, column=col).value

        if not cell_value:
            break

        val = str(cell_value)
        if val in {"Horizontal", "Vertical"}:
            orientations.add(val)

        row += 1

    if not orientations:
        return ""
    elif len(orientations) == 1:
        return orientations.pop()
    else:
        return ", ".join(sorted(orientations))


def get_max_row(xl):
    for i in range(1, 1000):
        if xl.cell(row=i, column=1).value == "Total":
            return i

    st.error("Total Row not found in uploaded spreasheet.")


def add_total_border(xl, row, col_start, col_end):

    top_border = Border(
        top=Side(style="thin", color="FD5300")
    )

    for col in range(col_start, col_end + 1):
        cell = xl.cell(row=row, column=col)

        cell.border = Border(
            top=top_border.top
        )

        adjust_cell(cell)

    return xl


def add_merged_cell(xl, row, col_start, col_end, text, bold=False):
    merge_cells(xl, row, row, col_start, col_end)
    total_cell = xl.cell(row=row, column=1)

    set_cell(total_cell, text, bold=bold, alignment="center")
    return xl


def add_total_rows(xl, row_max, col_max):

    ref_cell = xl.cell(row=row_max-1, column=col_max-1)

    cell = xl.cell(row=row_max, column=col_max)
    col_ref = f"{chr(64 + col_max)}"
    total_row = row_max
    set_cell(cell, f'=SUM(${col_ref}$2:INDIRECT("{col_ref}" & ROW()-1))')
    cell.number_format = ref_cell.number_format
    xl = add_merged_cell(xl, row_max, 1, col_max - 1, "Total")

    row_max += 1
    gst_cell = xl.cell(row=row_max, column=col_max)
    set_cell(gst_cell, "+ 18% GST", alignment="right")
    xl = add_merged_cell(xl, row_max, 1, col_max - 1, "")
    gst_cell.parent.row_dimensions[gst_cell.row].height = 30

    row_max += 1
    gst_total = xl.cell(row_max, column=col_max)
    set_cell(gst_total, f'=1.18*{col_ref}{total_row}', bold=True)
    gst_total.number_format = ref_cell.number_format
    xl = add_merged_cell(xl, row_max, 1, col_max - 1, "Total Project Value", bold=True)
    xl = add_total_border(xl, row_max, 1, col_max)

    return xl, row_max


def get_total_cols(xl, curr_row, start_col, end_col):
    vars = {}

    for col in range(start_col, end_col + 1):
        col_title = xl.cell(row=3, column=col).value
        if col_title is not None:
            col_total = xl.cell(row=curr_row, column=col).value
            vars[col_title] = col_total

    return vars


def get_cell_ref(col, row):
    return f"{chr(64 + col)}{row}"


def evaluate_formula(wb_obj, sheet_name, cell):
    compiler = ModelCompiler()
    new_model = compiler.read_and_parse_archive(wb_obj)
    evaluator = Evaluator(new_model)

    val = evaluator.evaluate(f"{sheet_name}!{cell}")
    return val


def set_cell_and_save(wb, val, row, col):
    xl = wb.worksheets[0]
    pitch_cell = xl.cell(row=row, column=col)
    set_cell(pitch_cell, val)

    temp_output = BytesIO()
    wb.save(temp_output)
    temp_output.seek(0)

    return temp_output


def adjust_cell(cell, threshold=35):

    text_length = len(str(cell.value))
    lines = 1
    if text_length > threshold:
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        lines = math.ceil(text_length / threshold)
        new_height = 19 * (lines + 1)
        cell.parent.row_dimensions[cell.row].height = new_height


def color_cells(xl, row_start, row_end, col_start, col_end, color):
    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    for r in xl.iter_rows(
        min_row=row_start, max_row=row_end, min_col=col_start, max_col=col_end
    ):

        for cell in r:
            cell.fill = fill


def generate_commercial_table(data):

    BASE_DIR = Path(__file__).resolve().parent
    path = BASE_DIR/"files"/"reference_xls"/"commercials.xlsx"
    commercial_wb = openpyxl.load_workbook(path)
    commercial_xl = commercial_wb.worksheets[0]
    curr_row = 2

    for i in range(len(data)):
        for idx in range(4):
            cell = commercial_xl.cell(row=curr_row, column=idx + 1)
            if idx == 0:
                set_cell(cell, data[i][idx], alignment="left")
            else:
                set_cell(cell, data[i][idx])
            adjust_cell(cell)

        total_cell = commercial_xl.cell(row=curr_row, column=5)
        set_cell(total_cell, f"=B{curr_row}*D{curr_row}")

        curr_row += 1

    commercial_xl, curr_row = add_total_rows(commercial_xl, curr_row, 5)

    commercial_xl.delete_rows(curr_row+1, commercial_xl.max_row)

    output = BytesIO()
    commercial_wb.save(output)
    output.seek(0)

    return output


def combine_commercial_xls(wb, dfs):

    xl = wb.worksheets[0]
    start = curr_row = 2

    for i in range(len(dfs)):
        df = dfs[i]
        curr_row = start
        for item in df:
            for idx in range(4):
                cell = xl.cell(row=curr_row, column=idx + 2)
                if idx == 0:
                    set_cell(cell, item[idx], alignment="left")
                    adjust_cell(cell)
                elif idx == 1:
                    set_cell(cell, round(item[idx], 0))
                elif idx == 2:
                    set_cell(cell, item[idx], alignment="center")
                else:
                    set_cell(cell, item[idx])

                thin_black = Side(style="hair", color="000000")
                all_thin_border = Border(
                    left=thin_black,
                    right=thin_black,
                    top=thin_black,
                    bottom=thin_black
                )

                cell.border = all_thin_border

            total_cell = xl.cell(row=curr_row, column=6)
            set_cell(total_cell, f"=C{curr_row}*E{curr_row}")

            curr_row += 1

        curr_row -= 1

        merge_cells(xl, start, curr_row, 1, 1)
        s_no = xl.cell(row=start, column=1)
        set_cell(s_no, i + 1)

        merge_cells(xl, start, curr_row, 7, 7)
        product_total = xl.cell(row=start, column=7)
        set_cell(product_total, f"=SUM(F{start}:F{curr_row})")

        color_cells(xl, start, curr_row, 1, 7, COLORS[i % 2])

        start = curr_row + 1

    xl, start = add_total_rows(xl, start, 7)

    xl.delete_rows(start+1, xl.max_row)

    return wb


def combine_xls(dfs):

    BASE_DIR = Path(__file__).resolve().parent
    path = BASE_DIR/"files"/"reference_xls"/"combined_commercials.xlsx"
    wb = openpyxl.load_workbook(path)

    wb = combine_commercial_xls(wb, dfs)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output
